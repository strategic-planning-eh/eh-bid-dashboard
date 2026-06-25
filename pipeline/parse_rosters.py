import openpyxl, json, unicodedata, re, statistics

def norm_ar(s):
    s=unicodedata.normalize('NFKC', str(s or '')).lower()
    for a,b in [('أ','ا'),('إ','ا'),('آ','ا'),('ة','ه'),('ى','ي'),('ـ','')]: s=s.replace(a,b)
    return s
def canon(s):
    t=norm_ar(s); t=re.sub(r'[^\w\u0600-\u06FF ]',' ',t); return ' '.join(t.split())
def is_eh(s):
    t=norm_ar(s)
    if 'environmental horizon' in t or 'afaq' in t: return True
    return ('افاق' in t and 'بيئ' in t)
def is_amt(v): return isinstance(v,(int,float)) and 1000<=v<1e11
DQ=['غير مطابق','مستبعد','لم يجتاز','غير متطابق','مستبعده','غير ﻣﻄﺎﺑﻖ']
def is_dq(v): return isinstance(v,str) and any(d in norm_ar(v) or d in str(v) for d in DQ)
HEADERW=['bidder name','اسم المورد','اسم الشركه','supplier','description','rfx','total amount','وصف','رقم','الرقم','remarks','status','date','اسم العارض','المورد','قيمه','الشركات','بيانات','م','no','sn','open','time','response','بند','اجمالي','vat','ضريبه','قائمه الموردين','تاريخ']
def is_name(v):
    if not isinstance(v,str): return False
    t=v.strip()
    if len(t)<6 or len(t)>90: return False
    if is_amt(v): return False
    cn=norm_ar(t)
    if cn in HEADERW or cn in ('م','no','sn'): return False
    for ph in ['الشركات الي قدمت','قائمه الموردين','الموردين المتقدمين','الموردين المرسى','المرسى عليه','تقديم خدمات','اعداد دراسه','تنفيذ اعمال','اجراء دراسه','قياس بيئي','عدد الشركات','الشركه الفائزه','تاريخ فتح','رقم المنافسه','تقرير فتح','فتح العروض','الموعد النهائي','تاريخ الترسيه']:
        if ph in cn: return False
    # reject project/service descriptions (start with these words, not company names)
    first=cn.split()[0] if cn.split() else ''
    if first in ['مشروع','خدمات','اعداد','تنفيذ','اجراء','توريد','صيانه','تطوير','دراسه','اعاده','تركيب','تقييم'] and 'شركه' not in cn and 'مؤسسه' not in cn:
        return False
    # must contain letters
    return bool(re.search(r'[A-Za-z\u0600-\u06FF]{3,}',t))

def parse(rs):
    maxr=min(rs.max_row,500); maxc=rs.max_column
    grid=[[rs.cell(r+1,c+1).value for c in range(maxc)] for r in range(maxr)]
    # name column: 1) header match ('Bidder Name'/'اسم المورد')  2) most name-like cells (excl description)
    DESC=['description','وصف','rfx','بيان المنافسه','specification']
    nc=None
    for c in range(maxc):
        for r in range(min(12,maxr)):
            h=norm_ar(grid[r][c])
            if any(k in h for k in ['bidder name','اسم المورد','اسم الشركه','supplier','اسم العارض','المورد']) and not any(d in h for d in DESC):
                nc=c; break
        if nc is not None: break
    if nc is None:
        nscore={c:sum(1 for row in grid if is_name(row[c]) and not any(b in norm_ar(row[c]) for b in ['قيمه','total amount']+DESC)) for c in range(maxc)}
        # exclude columns whose header marks them as description
        for c in list(nscore):
            for r in range(min(12,maxr)):
                if any(d in norm_ar(grid[r][c]) for d in DESC): nscore[c]=0; break
        if not nscore or max(nscore.values())<2: return None
        nc=max(nscore,key=nscore.get)
    # price column = header/spread heuristic (optional)
    cands=[]
    for c in range(maxc):
        vals=[row[c] for row in grid if is_amt(row[c])]
        if len(vals)>=2:
            cands.append((c,len(vals),max(vals)/max(min(vals),1)))
    pc=None
    if cands:
        def hscore(c):
            for r in range(min(12,maxr)):
                h=norm_ar(grid[r][c])
                if any(k in h for k in ['قيمه العرض','قيمه الترسيه','total amount','اجمالي','القيمه الماليه','amount','قيمه العطاء']): return 2
                if any(k in h for k in ['رقم','reference','no.','الرقم']): return -3
            return 0
        pc=max(cands, key=lambda x:(hscore(x[0])*10 + (5 if x[2]>1.5 else -2) + min(x[1],50)))[0]
    out=[]; seen=set()
    for row in grid:
        nm=row[nc]
        if not is_name(nm): continue
        k=canon(nm)[:40]
        if k in seen or not k: continue
        pv=row[pc] if pc is not None else None
        dq=is_dq(pv) if pc is not None else False
        amt=float(pv) if is_amt(pv) else None
        seen.add(k); out.append([str(nm).strip(), amt, dq, is_eh(nm)])
    return out if len(out)>=2 else None

if __name__=='__main__':
    allr={}
    for f,yr in [('bids2025.xlsx',2025),('bids2026.xlsx',2026)]:
        wb=openpyxl.load_workbook(f, data_only=True)
        for sh in [s for s in wb.sheetnames if s.startswith('Bid#')]:
            sn=sh.replace('Bid#','')
            r=parse(wb[sh])
            if r: allr[f'{yr}-{sn}']=r
    json.dump(allr, open('rosters_all.json','w'), ensure_ascii=False)
    npriced=sum(1 for v in allr.values() if sum(1 for x in v if x[1]) >=2)
    eh_rosters=sum(1 for v in allr.values() if any(x[3] for x in v))
    print('Total rosters (incl names-only):',len(allr),'| with >=2 prices:',npriced,'| with EH present:',eh_rosters)
