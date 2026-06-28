import openpyxl, re, json, datetime
def num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v) if not (isinstance(v,float) and v!=v) else None
    s=re.sub(r'[^\d.\-]','',str(v))
    try: return float(s) if s not in ('','-','.') else None
    except: return None
def dstr(v):
    if isinstance(v,datetime.datetime): return v.strftime('%Y-%m-%d')
    return None
def mkey(v):
    if isinstance(v,datetime.datetime): return v.strftime('%Y-%m')
    return None

def colmap(ws,hr):
    m={}
    for c in range(1,ws.max_column+1):
        h=str(ws.cell(hr,c).value or '').replace('\n',' ').strip()
        if not h: continue
        if h.startswith('SN#') and 'Year' not in h and 'sn' not in m: m['sn']=c
        elif 'Bid launch' in h: m['launch']=c
        elif 'Bid Title' in h: m['title']=c
        elif 'Reference No' in h: m['ref']=c
        elif 'Project Duration (Month)' in h: m['dur']=c
        elif 'Client/Proponent' in h: m['client']=c
        elif 'Platform' in h: m['platform']=c
        elif 'Submission Deadline' in h: m['submit']=c
        elif h.endswith('CEO'): m['ceo']=c
        elif 'Fainance' in h or 'Finance Dep' in h: m['fin']=c
        elif 'Tech. Dep. assigned' in h: m['svcdept']=c
        elif h.endswith('Tech. Dep.'): m['tech']=c
        elif 'BD Dep' in h: m['bd']=c
        elif 'PM Dep' in h: m['pm']=c
        elif 'Admin. Dep' in h: m['admin']=c
        elif 'Legal Affairs' in h: m['legal']=c
        elif 'Bid Committee decision' in h: m['committee']=c
        elif 'Comment' in h or 'ملاحظ' in h: m['comments']=c
        elif 'قيمة العرض المالي المقدم (بدون' in h: m['offer']=c
        elif 'Discount Request' in h: m['disc']=c
        elif 'قيمة الضمان البنكي' in h: m['bankval']=c
        elif h.startswith('عدد الشركات'): m['nbid']=c
        elif h.startswith('الشركات الي قدمت'): m['bidders']=c
        elif 'الشركة الفائزة' in h: m['winner']=c
        elif 'قيمة العرض  الفائز' in h or ('قيمة العرض' in h and 'فائز' in h): m['winval']=c
        elif h.startswith('حالة') and 'ترسية' in h: m['status']=c
    return m

EH=['افاق البيئة','آفاق البيئة','افاق البيئه','آفاق البيئه','environmental horizon']
def is_eh(s): s=str(s or '').lower(); return any(t in s for t in EH)
def map_status(raw):
    s=str(raw or '').strip()
    if not s: return None
    n=s.replace('ة','ه').replace('أ','ا').replace('إ','ا').replace('آ','ا').replace('ى','ي')
    if 'لم تتم' in n or 'لم يتم' in n: return 'Not awarded'   # لم تتم الترسية = EH lost
    if 'تمت الترسيه' in n or 'تم الترسيه' in n or 'ترسيت' in n or 'ترسيه علينا' in n or 'فزنا' in n or 'awarded' in n.lower(): return 'Awarded'  # تمت الترسية = EH won
    if 'الغاء' in n or 'ملغ' in n or 'cancel' in n.lower(): return 'Cancelled'
    if 'جاري' in n or 'قيد' in n or 'progress' in n.lower(): return 'In progress'
    return s
def acc(v): 
    s=str(v or '').strip().lower()
    return 'accept' in s or 'موافق' in s or s=='yes'

allbids=[]; rosters={}
for f,yr in [('bids2025.xlsx',2025),('bids2026.xlsx',2026)]:
    wb=openpyxl.load_workbook(f, data_only=True); ws=wb[wb.sheetnames[0]]
    hr=5 if yr==2025 else 4
    for ri in range(1,8):
        if str(ws.cell(ri,1).value or '').strip().startswith('SN'): hr=ri; break
    m=colmap(ws,hr)
    for ri in range(hr+1, ws.max_row+1):
        sn=ws.cell(ri,m['sn']).value
        if sn in (None,''): continue
        try: sn=int(float(sn))
        except: continue
        g=lambda key: ws.cell(ri,m[key]).value if key in m else None
        st=map_status(g('status'))
        winner=g('winner')
        # skip empty pre-numbered placeholder rows (SN present but no real content)
        _content=[g('title'),g('client'),g('launch'),g('platform'),g('offer'),winner,g('nbid'),g('committee'),g('dur'),g('svcdept'),g('bankval')]
        if not any(str(x).strip() for x in _content if x not in (None,'')): continue
        b=dict(year=yr, sn=sn,
            launch=dstr(g('launch')), month=mkey(g('launch')),
            submit=dstr(g('submit')), submonth=mkey(g('submit')),
            title=str(g('title') or '').strip(),
            client=str(g('client') or '').strip(),
            platform=str(g('platform') or '').strip(),
            dur=num(g('dur')),
            svcdept=str(g('svcdept') or '').strip(),
            committee=str(g('committee') or '').strip(),
            comments=str(g('comments') or '').strip(),
            offer=num(g('offer')),
            discount=str(g('disc') or '').strip(),
            bankval=num(g('bankval')),
            nbid=num(g('nbid')),
            winner=str(winner or '').strip(),
            winval=num(g('winval')),
            eh_won=(True if st=='Awarded' else False if st=='Not awarded' else (is_eh(winner) if winner else None)),
            decided=(st in ('Awarded','Not awarded')) or bool(winner and str(winner).strip()),
            status=st,
            appr=dict(ceo=acc(g('ceo')),fin=acc(g('fin')),tech=acc(g('tech')),bd=acc(g('bd')),pm=acc(g('pm')),admin=acc(g('admin')),legal=acc(g('legal'))),
        )
        allbids.append(b)
    # rosters from Bid# sheets
    for sh in wb.sheetnames:
        if not sh.startswith('Bid#'): continue
        try: sn=int(sh.replace('Bid#',''))
        except: continue
        rs=wb[sh]; maxc=rs.max_column
        hri=None
        for r in range(1,rs.max_row+1):
            txt=' '.join(str(rs.cell(r,c).value) for c in range(1,maxc+1) if rs.cell(r,c).value)
            if 'اسم المورد' in txt or 'Bidder Name' in txt or 'Total Amount' in txt: hri=r; break
        if not hri: continue
        namec=valc=None
        for c in range(1,maxc+1):
            h=str(rs.cell(hri,c).value or '')
            if 'اسم المورد' in h or 'Bidder Name' in h: namec=c
            if 'قيمة العرض' in h or 'Total Amount' in h or 'قيمة الترسية' in h: valc=c
        if not namec: continue
        out=[]
        for r in range(hri+1, rs.max_row+1):
            nm=rs.cell(r,namec).value
            if not nm: continue
            val=rs.cell(r,valc).value if valc else None
            dq=isinstance(val,str) and ('غير مطابق' in val)
            out.append([str(nm).strip(), None if dq else num(val), dq])
        if out: rosters[f'{yr}-{sn}']=out

json.dump({'bids':allbids,'rosters':rosters}, open('bidraw2.json','w'), ensure_ascii=False)
print('bids:',len(allbids),'| with launch date:',sum(1 for b in allbids if b['month']))
print('with platform:',sum(1 for b in allbids if b['platform']),'| with duration:',sum(1 for b in allbids if b['dur']))
print('with svc dept:',sum(1 for b in allbids if b['svcdept']),'| with committee:',sum(1 for b in allbids if b['committee']))
print('with offer:',sum(1 for b in allbids if b['offer']),'| with winner:',sum(1 for b in allbids if b['winner']),'| EH won:',sum(1 for b in allbids if b['eh_won']))
print('rosters:',len(rosters),'| roster line-items:',sum(len(v) for v in rosters.values()))
